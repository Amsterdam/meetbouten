import logging
from datetime import datetime
from typing import Any, Union

from django.contrib import admin
from django.db.models import Prefetch
from django.http import HttpResponse
from openpyxl import Workbook

from bouwblokken.models import Bouwblok, Controlepunt, Kringpunt, Referentiepunt
from metingen.models import Hoogtepunt, MetingHerzien

logger = logging.getLogger(__name__)


class BouwblokActionsMixin:
    @admin.action(description="Maak rapportage bouwblok")
    def get_report_bouwblok(self, request, queryset: list[Bouwblok]):
        """Make report about all selected bouwblokken"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Rapportage op basis van HERZIEN stelsel"])

        for bouwblok in queryset:
            data = self.fetch_bouwblok_data(bouwblok)
            ws = wb.create_sheet(f"Bouwblok {data['block']}")
            self.create_bouwblok_sheet(data, ws)

        filename = f"rapportage-bouwblok-{datetime.today().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response

    def fetch_bouwblok_data(self, bouwblok: Bouwblok) -> dict[str, Any]:
        # Prefetch "status" from prefetched "hoogtepunt"
        return {
            "block": bouwblok.nummer,
            "aansluitpunt": bouwblok.aansluitpunt_id,
            "controlepunt": bouwblok.controlepunt_id,
            "referentiepunten": self._get_punten_data(
                Referentiepunt.objects.filter(bouwblok=bouwblok).prefetch_related(
                    Prefetch(
                        "hoogtepunt",
                        queryset=Hoogtepunt.objects.prefetch_related("status"),
                    )
                )
            ),
            "controlepunten": self._get_punten_data(
                Controlepunt.objects.filter(bouwblok=bouwblok).prefetch_related(
                    Prefetch(
                        "hoogtepunt",
                        queryset=Hoogtepunt.objects.prefetch_related("status"),
                    )
                )
            ),
            "kringpunten": self._get_punten_data(
                Kringpunt.objects.filter(bouwblok=bouwblok).prefetch_related(
                    Prefetch(
                        "hoogtepunt",
                        queryset=Hoogtepunt.objects.prefetch_related("status"),
                    )
                )
            ),
        }

    def _get_punten_data(
        self, punten: Union[list[Referentiepunt], list[Controlepunt], list[Kringpunt]]
    ) -> list[dict[str, Any]]:
        data = []
        for punt in punten:
            hoogtepunt = punt.hoogtepunt
            if hoogtepunt.status.omschrijving == "Vervallen":
                logger.warning(
                    f"Vervallen hoogtepunt {hoogtepunt} in bouwblok {punt.bouwblok}"
                )
                continue
            last_meting = (
                MetingHerzien.objects.filter(hoogtepunt=hoogtepunt)
                .order_by("inwindatum")
                .last()
            )
            data.append(
                dict(
                    Nummer=punt.id,
                    Omschrijving=hoogtepunt.omschrijving,
                    Merk=hoogtepunt.merk_id,
                    Xmuur=hoogtepunt.xmuur,
                    Ymuur=hoogtepunt.ymuur,
                    Windrichting=hoogtepunt.windr,
                    Hoogte=last_meting.hoogte if last_meting else None,
                    Datum=last_meting.inwindatum if last_meting else None,
                    Hoogtepunt=hoogtepunt.nummer,
                )
            )
        return data

    def create_bouwblok_sheet(self, data, ws):
        ws.append(["Bouwblok", data["block"]])
        ws.append([""])
        ws.append(["Aansl.punt blok", data["aansluitpunt"]])
        ws.append(["Contr.punt blok", data["controlepunt"]])
        ws.append([""])

        ws.append(
            [
                "Nummer",
                "Omschrijving",
                "Merk",
                "Xmuur",
                "Ymuur",
                "Windrichting",
                "Hoogte",
                "Datum",
                "Hoogtepunt",
            ]
        )
        ws.append(["Referentiepunten:"])
        if len(data["referentiepunten"]) < 2:
            ws.append(
                [
                    f"BOUWBLOK {data['block']} HEEFT TE WEINIG OF HEEFT VERVALLEN REFERENTIEPUNTEN"
                ]
            )
        [ws.append(list(row.values())) for row in data["referentiepunten"]]
        ws.append([""])
        ws.append(["Controlepunten:"])
        if len(data["controlepunten"]) == 0:
            ws.append([f"BOUWBLOK {data['block']} HEEFT GEEN CONTROLEPUNTEN"])
        [ws.append(list(row.values())) for row in data["controlepunten"]]
        ws.append([""])
        ws.append(["Kringpunten:"])
        if len(data["kringpunten"]) == 0:
            ws.append([f"BOUWBLOK {data['block']} HEEFT GEEN KRINGPUNTEN"])
        [ws.append(list(row.values())) for row in data["kringpunten"]]

    @admin.action(description="Maak rapportage historische data")
    def get_report_history(self, request, queryset):
        filename = f"rapportage-historie-{datetime.today().strftime('%Y-%m-%d')}.xlsx"
        field_names = [
            "Bouwblok nummer",
            "Meting nummer",
            "Hoogtepunt nummer",
            "Inwindatum",
            "Hoogte",
            "Verschil vorig",
            "Verschil nulmeting",
        ]
        data = self.collect_data(queryset)

        wb = Workbook()
        ws = wb.active
        ws.append(field_names)
        [ws.append(row) for row in data]

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response

    def collect_data(self, queryset):
        data = []
        for bouwblok in queryset:
            hoogtepunten = Hoogtepunt.objects.filter(
                id__in=Kringpunt.objects.filter(bouwblok=bouwblok).values("hoogtepunt")
            ).order_by("nummer")
            for hoogtepunt in hoogtepunten:
                metingen = MetingHerzien.objects.filter(
                    hoogtepunt=hoogtepunt.id
                ).order_by("inwindatum")
                start_meting = metingen.first()
                if not start_meting:
                    continue
                start_hoogte = start_meting.hoogte
                laatste_hoogte = start_hoogte
                for meting in metingen:
                    data.append(
                        [
                            bouwblok.nummer,
                            meting.id,
                            hoogtepunt.nummer,
                            meting.inwindatum,
                            round(meting.hoogte, 4),
                            round(meting.hoogte - laatste_hoogte, 4),
                            round(meting.hoogte - start_hoogte, 4),
                        ]
                    )
                    laatste_hoogte = meting.hoogte
        return data
